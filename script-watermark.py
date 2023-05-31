# -------------------- History ---------------------------
#
# Adding DS/DZ logo
#
# Created by Jirameth Kaewsuwan
# Email meth.jirameth@gmail.com
#
# Created on 13:50PM 07th August 2021
# Version 1.0 
#
# Copy right by Jirameth Kaewsuwan
#
# Change history
# (1) 07/08/2021 -> Created program including detecting error and log file
# (2) 09/08/2021 -> Adding create directory output function
#
# --------------------------------------------------------


# import all library -------------------------------------
import cv2
import time 
import pyautogui
import logging
from openpyxl import load_workbook
import datetime
import os
import shutil
# --------------------------------------------------------

# initial string name ------------------------------------

picture_folder  = 'HMI-photo/'
logo_folder = 'DS-DZ/'
output_folder = 'Output/'
jpg_type = '.jpg'
ds_logo = 'DS-sign'
dz_logo = 'DZ-sign'

# --------------------------------------------------------



# Date information ---------------------------------------

current_date = datetime.datetime.now()
current_time = 	str(current_date.strftime("%Y")) + \
	       		str(current_date.strftime("%m")) + \
	       		str(current_date.strftime("%d")) + '-' +\
           		str(current_date.strftime("%H")) + \
           		str(current_date.strftime("%M")) + \
           		str(current_date.strftime("%S"))

# --------------------------------------------------------

# Create folder output -----------------------------------

output_folder_num = 'HMI-' + current_time +'/'
directory_output = output_folder + output_folder_num
os.mkdir(directory_output)
# --------------------------------------------------------

# Logging system -----------------------------------------

#Create log name
log_name = current_time + '.log'

#log pattern
logging.basicConfig(filename=directory_output + '/' + log_name, level=logging.INFO, format='%(asctime)s - %(levelname)s ->	%(message)s')
logging.basicConfig(filename=directory_output + '/' + log_name, level=logging.WARNING, format='%(asctime)s - %(levelname)s ->	%(message)s')
# --------------------------------------------------------

# Stop running system ------------------------------------
def stop_running() :
	log_text = 'Running script finished'
	print(log_text)
	logging.info(log_text)
	raise SystemExit	#Stop running
# --------------------------------------------------------

# Notification -------------------------------------------
def notification(str_input = '-') :
	print(str_input)
	logging.info(str_input)
# --------------------------------------------------------

# Excel --------------------------------------------------
def read_excel_data(filename = 'SC-list - Copy.xlsx') :
	try :
		#load workbook
		workbook = load_workbook(filename)	

	except FileNotFoundError as error_type :	#Could find workbook
		log_text = 'Handing error - ' + str(error_type) + '. Could not find excel name ' + filename + '.'
		notification(log_text)
		pyautogui.confirm(text='You must put excel name SC-list.xlsx on the location.', title='Error', buttons=['OK'])
		stop_running()

	except PermissionError as error_type :		#Work book is being opened
		log_text = 'Handing error - ' + str(error_type) + '. Excel name ' + filename + ' is being opened. Could not open it.'
		notification(log_text)
		pyautogui.confirm(text='You must close excel before running program or set permission first.', title='Error', buttons=['OK'])
		stop_running()

	#activate workbook
	sheet = workbook.active

	log_text = 'Excel filename : ' + str(filename) + 'has been opened.'
	notification(log_text)

	#output value from excel
	values = []

	for row in sheet.iter_rows(min_row=2, values_only=True):	#loop to get data
		values.append(row)	#Add to list

	log_text = 'Data have been gotten (' + str(len(values)) + ' lists).'
	notification(log_text)

	return values 	#return to function
# --------------------------------------------------------

# Create DS/DZ picture --------------------------------------
def adding_logo(picture_path, logo_path, output_path, SC, picture_name, station = 'empty'):

	# Position initail
	start_point_y = 20
	start_point_x = 20

	# Read picture
	logo = cv2.imread(str(logo_path))
	img = cv2.imread(str(picture_path))

	
	try :
		# Check size of photo
		h_img, w_img, _ = img.shape
		h_logo, w_logo, _ = logo.shape

	except AttributeError as error_type :	# Not found picture   -   #Couldn't check size because couldn't find picture
		log_text = 'Handing error - ' + str(error_type) + '. Could not find photo on this path (' + picture_path + ')'
		notification(log_text)
		pyautogui.confirm(text='Please input all DS/DZ photos.', title='Error', buttons=['OK'])

	try :
		# Calculate adding logo position
		top_y = start_point_y
		bottom_y = start_point_y + h_logo
		left_x = start_point_x
		right_x = start_point_x + w_logo

		# Adding logo
		img[top_y: bottom_y, left_x: right_x] = logo

		# Save photo
		cv2.imwrite(output_path,img)

		log_text = 'st.' + station + ' - picture name : ' + str(picture_name) + ' has been added ' + SC + ' logo.'
		notification(log_text)

	except :
		log_text = 'st.' + station + ' - picture name : ' + str(picture_name) + ' is not found'
		notification(log_text)

# -------------------------------------------------------


# Main ---------------------------------------------------

# Reading excel data
excel_data_list = read_excel_data()


for check_process in excel_data_list : 	# Loop all photo name 

	#Excel list to variable
	picture_name = str(check_process[0])
	SC_symbol = str(check_process[1])
	sub_pic_folder = str(check_process[2]) +'/'
	station = str(check_process[2])

	#Check folder and create folder
	check_directory = output_folder + output_folder_num + sub_pic_folder	#Directory output path
	try :
		os.listdir(check_directory)		#Check directory 
	except FileNotFoundError :
		os.mkdir(check_directory)		#Create directory
		log_text = 'folder location : ' + os.getcwd() + '/' + str(check_directory) + ' was created'
		notification(log_text)

	#Create path
	picture_path = picture_folder + sub_pic_folder + picture_name + jpg_type
	output_path = output_folder + output_folder_num + sub_pic_folder + picture_name + jpg_type

	#Check SC
	if SC_symbol == 'DS':
		logo_path = logo_folder + ds_logo + jpg_type	#logo path
		adding_logo(picture_path,logo_path,output_path,SC_symbol,picture_name,station)

	elif SC_symbol =='DZ' :
		logo_path = logo_folder + dz_logo + jpg_type	#logo path
		adding_logo(picture_path,logo_path,output_path,SC_symbol,picture_name,station)

	elif SC_symbol =='-' :
		log_text = 'picture name : ' + str(picture_name) + ' is not DS/DZ.'
		notification(log_text)

	else :
		log_text = 'picture name : ' + str(picture_name) +  ' is error! Type of picture is mismatch.'
		notification(log_text)
		pyautogui.confirm(text=log_text, title='Error', buttons=['OK'])

#Finish running
log_text = 'Running script finished'
notification(log_text)

# -------------------------------------------------------

		
	





